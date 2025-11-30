"""
Email Export Script for Cron Job
Exports user signups and intern applications to Excel files
"""

import os
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Database configuration
DATABASE_URL = os.getenv('DATABASE_URL')
USE_POSTGRES = DATABASE_URL is not None

def get_db_connection():
    """Get database connection"""
    if USE_POSTGRES:
        import psycopg2
        return psycopg2.connect(DATABASE_URL)
    else:
        import sqlite3
        return sqlite3.connect('aisolutions.db')

def create_excel_with_style(filename, headers, data):
    """Create a styled Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(filename)
    print(f"✅ Created: {filename}")

def export_user_signups():
    """Export all user signup emails"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all users
        cursor.execute('''
            SELECT name, email, phone, address, created_at
            FROM users
            ORDER BY created_at DESC
        ''')
        
        users = cursor.fetchall()
        conn.close()
        
        if not users:
            print("⚠️ No users found")
            return
        
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'exports/user_signups_{timestamp}.xlsx'
        
        headers = ['Full Name', 'Email', 'Phone', 'Address', 'Signup Date']
        data = []
        
        for user in users:
            data.append([
                user[0],  # name
                user[1],  # email
                user[2],  # phone
                user[3],  # address
                str(user[4]) if user[4] else 'N/A'  # created_at
            ])
        
        create_excel_with_style(filename, headers, data)
        print(f"📧 Exported {len(users)} user signups")
        
    except Exception as e:
        print(f"❌ Error exporting user signups: {e}")
        import traceback
        traceback.print_exc()

def export_intern_applications():
    """Export all intern application emails with job titles"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Get all applications
        cursor.execute('''
            SELECT full_name, email, phone, position, college, degree, 
                   semester, year, status, applied_at, linkedin, github
            FROM applications
            ORDER BY applied_at DESC
        ''')
        
        applications = cursor.fetchall()
        conn.close()
        
        if not applications:
            print("⚠️ No applications found")
            return
        
        # Create Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'exports/intern_applications_{timestamp}.xlsx'
        
        headers = ['Full Name', 'Email', 'Phone', 'Job Title/Position', 'College', 
                   'Degree', 'Semester', 'Year', 'Status', 'Applied Date', 'LinkedIn', 'GitHub']
        data = []
        
        for app in applications:
            data.append([
                app[0],   # full_name
                app[1],   # email
                app[2],   # phone
                app[3],   # position (job title)
                app[4],   # college
                app[5],   # degree
                app[6],   # semester
                app[7],   # year
                app[8],   # status
                str(app[9]) if app[9] else 'N/A',  # applied_at
                app[10] or 'N/A',  # linkedin
                app[11] or 'N/A'   # github
            ])
        
        create_excel_with_style(filename, headers, data)
        print(f"📧 Exported {len(applications)} intern applications")
        
    except Exception as e:
        print(f"❌ Error exporting applications: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main export function"""
    print("\n" + "="*60)
    print(f"🔄 Email Export Job Started - {datetime.now()}")
    print("="*60)
    
    # Create exports directory if it doesn't exist
    os.makedirs('exports', exist_ok=True)
    
    # Export user signups
    export_user_signups()
    
    # Export intern applications
    export_intern_applications()
    
    print("="*60)
    print("✅ Email Export Job Completed")
    print("="*60 + "\n")

if __name__ == '__main__':
    main()
